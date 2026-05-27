#!/usr/bin/env python3
"""Валидация заполненного шаблона справочников перед импортом в БД."""
import sys
from openpyxl import load_workbook

FILE = sys.argv[1] if len(sys.argv) > 1 else "Справочники_GrowFood.xlsx"
wb = load_workbook(FILE, data_only=True)

SHEET = {
    "vertical": "1. Вертикали",
    "location": "2. Локации",
    "vehicleType": "3. Типы ТС",
    "carrier": "4. Перевозчики",
    "customer": "5. Контрагенты",
    "vehicle": "6. Транспорт",
    "driver": "7. Водители",
    "route": "8. Маршруты",
}

def rows(sheet_name):
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        out.append({headers[i]: r[i] for i in range(len(headers)) if i < len(r)})
    return out

def s(v):
    return v.strip() if isinstance(v, str) else (None if v is None else str(v))

data = {k: rows(v) for k, v in SHEET.items()}
codes = {
    "vertical": {s(x.get("code")) for x in data["vertical"]},
    "location": {s(x.get("code")) for x in data["location"]},
    "vehicleType": {s(x.get("code")) for x in data["vehicleType"]},
    "carrier": {s(x.get("code")) for x in data["carrier"]},
    "customer": {s(x.get("code")) for x in data["customer"]},
}

errors = []
warns = []
def req(sheet, i, row, field):
    if s(row.get(field)) in (None, ""):
        errors.append(f"[{sheet}] строка {i}: пустое обязательное поле '{field}'")
def enum(sheet, i, row, field, allowed, required=True):
    v = s(row.get(field))
    if v in (None, ""):
        if required: errors.append(f"[{sheet}] строка {i}: пустое '{field}'")
        return
    if v not in allowed:
        errors.append(f"[{sheet}] строка {i}: '{field}'='{v}' не из {sorted(allowed)}")
def fk(sheet, i, row, field, target, required=True):
    v = s(row.get(field))
    if v in (None, ""):
        if required: errors.append(f"[{sheet}] строка {i}: пустая ссылка '{field}'")
        return
    if v not in codes[target]:
        errors.append(f"[{sheet}] строка {i}: '{field}'='{v}' нет в справочнике «{target}»")
def uniq(sheet, key):
    seen = {}
    for i, row in enumerate(data[sheet], start=2):
        v = s(row.get(key))
        if v in (None, ""): continue
        if v in seen: errors.append(f"[{sheet}] дубль '{key}'='{v}' (строки {seen[v]} и {i})")
        else: seen[v] = i

# --- Вертикали ---
uniq("vertical", "code")
for i, r in enumerate(data["vertical"], 2):
    req("vertical", i, r, "code"); req("vertical", i, r, "name")
    enum("vertical", i, r, "type", {"INTERNAL", "EXTERNAL"})
# --- Локации ---
uniq("location", "code")
for i, r in enumerate(data["location"], 2):
    req("location", i, r, "code"); req("location", i, r, "name")
    enum("location", i, r, "type", {"WAREHOUSE", "HUB", "KITCHEN", "DC", "RETAIL_POINT", "FACTORY"})
    enum("location", i, r, "ownerType", {"OWN", "CUSTOMER", "PARTNER"})
# --- Типы ТС ---
uniq("vehicleType", "code")
for i, r in enumerate(data["vehicleType"], 2):
    req("vehicleType", i, r, "code"); req("vehicleType", i, r, "name")
# --- Перевозчики ---
uniq("carrier", "code")
for i, r in enumerate(data["carrier"], 2):
    req("carrier", i, r, "code"); req("carrier", i, r, "name")
# --- Контрагенты ---
uniq("customer", "code")
for i, r in enumerate(data["customer"], 2):
    req("customer", i, r, "code"); req("customer", i, r, "name")
    fk("customer", i, r, "verticalCode", "vertical")
    enum("customer", i, r, "customerType", {"INTERNAL", "RETAIL_CHAIN", "EXTERNAL_COMPANY"})
    enum("customer", i, r, "partyRole", {"SHIPPER", "CONSIGNEE", "BOTH"}, required=False)
# --- Транспорт ---
uniq("vehicle", "plateNumber")
for i, r in enumerate(data["vehicle"], 2):
    req("vehicle", i, r, "plateNumber")
    fk("vehicle", i, r, "vehicleTypeCode", "vehicleType")
    fk("vehicle", i, r, "carrierCode", "carrier", required=False)
# --- Водители ---
for i, r in enumerate(data["driver"], 2):
    req("driver", i, r, "fullName")
    fk("driver", i, r, "carrierCode", "carrier", required=False)
# --- Маршруты ---
uniq("route", "code")
for i, r in enumerate(data["route"], 2):
    req("route", i, r, "code")
    fk("route", i, r, "originCode", "location")
    fk("route", i, r, "destinationCode", "location")
    enum("route", i, r, "routeType", {"DIRECT", "HUB", "MILK_RUN"})

print("=== Кол-во строк ===")
for k, v in SHEET.items():
    print(f"  {v}: {len(data[k])}")
print("\n=== Результат ===")
if errors:
    print(f"ОШИБКИ ({len(errors)}):")
    for e in errors: print("  ✗", e)
else:
    print("✓ Ошибок не найдено — данные согласованы, можно импортировать.")
if warns:
    print("\nПредупреждения:")
    for w in warns: print("  !", w)
