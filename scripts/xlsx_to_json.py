#!/usr/bin/env python3
"""Читает заполненный шаблон, нормализует значения по решениям и пишет /tmp/references.json."""
import sys, json
from openpyxl import load_workbook

FILE = sys.argv[1]
OUT = sys.argv[2] if len(sys.argv) > 2 else "/tmp/references.json"
wb = load_workbook(FILE, data_only=True)

def rows(name):
    ws = wb[name]; h = [c.value for c in ws[1]]; out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in r): continue
        out.append({h[i]: r[i] for i in range(len(h)) if i < len(r)})
    return out

def s(v):
    if v is None: return None
    v = str(v).strip()
    return v or None

def b(v, default=True):
    if v is None or (isinstance(v, str) and not v.strip()): return default
    if isinstance(v, bool): return v
    return str(v).strip().upper() in ("TRUE", "1", "ДА", "YES")

def num(v):
    if v in (None, "") or (isinstance(v, str) and not v.strip()): return None
    try: return float(v)
    except Exception: return None

def i(v):
    n = num(v)
    return int(n) if n is not None else None

LOC_TYPE = {"warehouse": "WAREHOUSE", "dc": "DC", "crossdock": "DC", "hub": "HUB",
            "kitchen": "KITCHEN", "retail_point": "RETAIL_POINT", "factory": "FACTORY"}
OWNER = {"own": "OWN", "client": "CUSTOMER", "customer": "CUSTOMER", "partner": "PARTNER"}
ROUTE_TYPE = {"ftl": "DIRECT", "direct": "DIRECT", "hub": "HUB", "milk_run": "MILK_RUN", "ltl": "HUB"}

def m(d, v, field):
    k = (s(v) or "").lower()
    if k in d: return d[k]
    raise SystemExit(f"Не знаю как сопоставить {field}='{v}'")

verticals = [{
    "code": s(r.get("code")), "name": s(r.get("name")),
    "type": "EXTERNAL" if (s(r.get("code")) or "").upper().startswith("LAAS") else "INTERNAL",
    "isActive": b(r.get("isActive")),
} for r in rows("1. Вертикали")]

locations = [{
    "code": s(r.get("code")), "name": s(r.get("name")),
    "type": m(LOC_TYPE, r.get("type"), "location.type"),
    "ownerType": m(OWNER, r.get("ownerType"), "location.ownerType"),
    "city": s(r.get("city")), "region": s(r.get("region")), "address": s(r.get("address")),
    "lat": num(r.get("lat")), "lon": num(r.get("lon")), "isActive": b(r.get("isActive")),
} for r in rows("2. Локации")]

vehicleTypes = [{
    "code": s(r.get("code")), "name": s(r.get("name")),
    "capacityKg": num(r.get("capacityKg")), "capacityPallets": i(r.get("capacityPallets")),
    "isRefrigerator": b(r.get("isRefrigerator"), default=False),
} for r in rows("3. Типы ТС")]

carriers = [{
    "code": s(r.get("code")), "name": s(r.get("name")), "inn": s(r.get("inn")), "kpp": s(r.get("kpp")),
    "contactPerson": s(r.get("contactPerson")), "phone": s(r.get("phone")), "email": s(r.get("email")),
    "isActive": b(r.get("isActive")), "notes": s(r.get("notes")),
} for r in rows("4. Перевозчики")]

customers = [{
    "code": s(r.get("code")), "name": s(r.get("name")), "inn": s(r.get("inn")), "kpp": s(r.get("kpp")),
    "fullLegalName": s(r.get("fullLegalName")),
    "verticalCode": "LAAS-LTL",          # решение: всем LAAS-LTL
    "customerType": "EXTERNAL_COMPANY",  # решение
    "partyRole": "BOTH",                 # решение
    "contactPerson": s(r.get("contactPerson")), "phone": s(r.get("phone")), "email": s(r.get("email")),
    "isActive": b(r.get("isActive")), "notes": s(r.get("notes")),
} for r in rows("5. Контрагенты")]

vehicles = [{
    "plateNumber": s(r.get("plateNumber")), "brandModel": s(r.get("brandModel")),
    "vehicleTypeCode": s(r.get("vehicleTypeCode")), "carrierCode": s(r.get("carrierCode")),
    "isActive": b(r.get("isActive")),
} for r in rows("6. Транспорт")]

drivers = [{
    "fullName": s(r.get("fullName")), "phone": s(r.get("phone")),
    "licenseNumber": s(r.get("licenseNumber")), "carrierCode": s(r.get("carrierCode")),
    "isActive": b(r.get("isActive")),
} for r in rows("7. Водители")]

routes = [{
    "code": s(r.get("code")), "name": s(r.get("name")),
    "originCode": s(r.get("originCode")), "destinationCode": s(r.get("destinationCode")),
    "distanceKm": num(r.get("distanceKm")), "estimatedHours": num(r.get("estimatedHours")),
    "routeType": m(ROUTE_TYPE, r.get("routeType"), "route.routeType"), "isActive": b(r.get("isActive")),
} for r in rows("8. Маршруты")]

out = {"verticals": verticals, "locations": locations, "vehicleTypes": vehicleTypes,
       "carriers": carriers, "customers": customers, "vehicles": vehicles,
       "drivers": drivers, "routes": routes}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=1)
print("OK ->", OUT)
for k, v in out.items():
    print(f"  {k}: {len(v)}")
