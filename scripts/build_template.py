#!/usr/bin/env python3
"""Генерирует Excel-шаблон справочников GrowFood Magistral для ручного заполнения."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_REQ = PatternFill("solid", fgColor="1F4E78")   # обязательные — тёмно-синий
HDR_OPT = PatternFill("solid", fgColor="808080")    # опциональные — серый
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MAXROW = 500  # на сколько строк раздаём валидацию/стиль

# col = (имя_поля, обязательно?, тип, допустимые значения|None, подсказка)
SHEETS = [
    ("1. Вертикали", "Vertical", [
        ("code", True, "текст, уникальный", None, "Код вертикали, напр. GROWFOOD, LAAS, RETAIL"),
        ("name", True, "текст", None, "Название"),
        ("type", True, "список", ["INTERNAL", "EXTERNAL"], "INTERNAL — внутренняя, EXTERNAL — внешняя (LaaS)"),
        ("isActive", False, "TRUE/FALSE", ["TRUE", "FALSE"], "Активна. По умолчанию TRUE"),
    ]),
    ("2. Локации", "Location", [
        ("code", True, "текст, уникальный", None, "Код локации, напр. MSK_DC"),
        ("name", True, "текст", None, "Название"),
        ("type", True, "список", ["WAREHOUSE", "HUB", "KITCHEN", "DC", "RETAIL_POINT", "FACTORY"], "Тип точки"),
        ("ownerType", True, "список", ["OWN", "CUSTOMER", "PARTNER"], "Чья точка"),
        ("city", False, "текст", None, "Город"),
        ("region", False, "текст", None, "Регион"),
        ("address", False, "текст", None, "Адрес"),
        ("lat", False, "число", None, "Широта, напр. 55.751244"),
        ("lon", False, "число", None, "Долгота, напр. 37.618423"),
        ("isActive", False, "TRUE/FALSE", ["TRUE", "FALSE"], "По умолчанию TRUE"),
    ]),
    ("3. Типы ТС", "VehicleType", [
        ("code", True, "текст, уникальный", None, "Код типа, напр. TRUCK_20T, REF_5T"),
        ("name", True, "текст", None, "Название, напр. Фура 20т"),
        ("capacityKg", False, "число", None, "Грузоподъёмность, кг"),
        ("capacityPallets", False, "целое", None, "Вместимость, паллет"),
        ("isRefrigerator", False, "TRUE/FALSE", ["TRUE", "FALSE"], "Рефрижератор. По умолчанию FALSE"),
    ]),
    ("4. Перевозчики", "Carrier", [
        ("code", True, "текст, уникальный", None, "Код перевозчика, напр. DELLIN"),
        ("name", True, "текст", None, "Название"),
        ("inn", False, "текст", None, "ИНН"),
        ("kpp", False, "текст", None, "КПП"),
        ("contactPerson", False, "текст", None, "Контактное лицо"),
        ("phone", False, "текст", None, "Телефон"),
        ("email", False, "текст", None, "Email"),
        ("isActive", False, "TRUE/FALSE", ["TRUE", "FALSE"], "По умолчанию TRUE"),
        ("notes", False, "текст", None, "Примечания"),
    ]),
    ("5. Контрагенты", "Customer", [
        ("code", True, "текст, уникальный", None, "Код контрагента, напр. MAGNIT"),
        ("name", True, "текст", None, "Название"),
        ("inn", False, "текст, уникальный", None, "ИНН (если указан — уникален)"),
        ("kpp", False, "текст", None, "КПП"),
        ("fullLegalName", False, "текст", None, "Полное юр. наименование"),
        ("verticalCode", True, "FK → Вертикали.code", None, "Код вертикали из листа «1. Вертикали»"),
        ("customerType", True, "список", ["INTERNAL", "RETAIL_CHAIN", "EXTERNAL_COMPANY"], "Тип контрагента"),
        ("partyRole", False, "список", ["SHIPPER", "CONSIGNEE", "BOTH"], "Роль. По умолчанию BOTH"),
        ("contactPerson", False, "текст", None, "Контактное лицо"),
        ("phone", False, "текст", None, "Телефон"),
        ("email", False, "текст", None, "Email"),
        ("isActive", False, "TRUE/FALSE", ["TRUE", "FALSE"], "По умолчанию TRUE"),
        ("notes", False, "текст", None, "Примечания"),
    ]),
    ("6. Транспорт", "Vehicle", [
        ("plateNumber", True, "текст, уникальный", None, "Гос. номер, напр. А123БВ78"),
        ("brandModel", False, "текст", None, "Марка/модель, напр. Volvo FH"),
        ("vehicleTypeCode", True, "FK → Типы ТС.code", None, "Код типа из листа «3. Типы ТС»"),
        ("carrierCode", False, "FK → Перевозчики.code", None, "Код перевозчика из листа «4. Перевозчики»"),
        ("isActive", False, "TRUE/FALSE", ["TRUE", "FALSE"], "По умолчанию TRUE"),
    ]),
    ("7. Водители", "Driver", [
        ("fullName", True, "текст", None, "ФИО"),
        ("phone", False, "текст", None, "Телефон"),
        ("licenseNumber", False, "текст", None, "Номер ВУ"),
        ("carrierCode", False, "FK → Перевозчики.code", None, "Код перевозчика из листа «4. Перевозчики»"),
        ("isActive", False, "TRUE/FALSE", ["TRUE", "FALSE"], "По умолчанию TRUE"),
    ]),
    ("8. Маршруты", "Route", [
        ("code", True, "текст, уникальный", None, "Код маршрута, напр. KLP-MSK"),
        ("name", False, "текст", None, "Название, напр. Колпино → Москва"),
        ("originCode", True, "FK → Локации.code", None, "Код локации-отправления (лист «2. Локации»)"),
        ("destinationCode", True, "FK → Локации.code", None, "Код локации-назначения (лист «2. Локации»)"),
        ("distanceKm", False, "число", None, "Расстояние, км"),
        ("estimatedHours", False, "число", None, "Плановое время в пути, ч"),
        ("routeType", True, "список", ["DIRECT", "HUB", "MILK_RUN"], "Тип маршрута"),
        ("isActive", False, "TRUE/FALSE", ["TRUE", "FALSE"], "По умолчанию TRUE"),
    ]),
    ("9. Договоры заказчиков", "CustomerContract", [
        ("contractNumber", True, "текст", None, "Номер договора"),
        ("customerCode", True, "FK → Контрагенты.code", None, "Код контрагента (лист «5. Контрагенты»)"),
        ("contractType", True, "список", ["LAAS_SERVICE", "RETAIL_SUPPLY", "INTERNAL_AGREEMENT"], "Тип договора"),
        ("validFrom", True, "дата ГГГГ-ММ-ДД", None, "Действует с, напр. 2026-01-01"),
        ("validTo", False, "дата ГГГГ-ММ-ДД", None, "Действует по (пусто = бессрочно)"),
        ("paymentTerms", False, "текст", None, "Условия оплаты"),
        ("notes", False, "текст", None, "Примечания"),
        ("isActive", False, "TRUE/FALSE", ["TRUE", "FALSE"], "По умолчанию TRUE"),
    ]),
    ("10. Договоры перевозчиков", "CarrierContract", [
        ("contractNumber", True, "текст", None, "Номер договора"),
        ("carrierCode", True, "FK → Перевозчики.code", None, "Код перевозчика (лист «4. Перевозчики»)"),
        ("validFrom", True, "дата ГГГГ-ММ-ДД", None, "Действует с"),
        ("validTo", False, "дата ГГГГ-ММ-ДД", None, "Действует по (пусто = бессрочно)"),
        ("paymentTerms", False, "текст", None, "Условия оплаты"),
        ("notes", False, "текст", None, "Примечания"),
        ("isActive", False, "TRUE/FALSE", ["TRUE", "FALSE"], "По умолчанию TRUE"),
    ]),
    ("11. Тарифы", "Tariff", [
        ("customerContractNumber", False, "FK → Договоры заказчиков", None, "Заполнить ЛИБО это, ЛИБО carrierContractNumber — не оба"),
        ("carrierContractNumber", False, "FK → Договоры перевозчиков", None, "Заполнить ЛИБО это, ЛИБО customerContractNumber — не оба"),
        ("routeCode", False, "FK → Маршруты.code", None, "Код маршрута (пусто = для любого)"),
        ("vehicleTypeCode", True, "FK → Типы ТС.code", None, "Код типа ТС (лист «3. Типы ТС»)"),
        ("pricePerTrip", False, "число", None, "Цена за рейс, ₽"),
        ("pricePerPallet", False, "число", None, "Цена за паллету, ₽"),
        ("pricePerKm", False, "число", None, "Цена за км, ₽"),
        ("validFrom", True, "дата ГГГГ-ММ-ДД", None, "Действует с"),
        ("validTo", False, "дата ГГГГ-ММ-ДД", None, "Действует по"),
        ("notes", False, "текст", None, "Примечания"),
    ]),
    ("12. Рыночные цены", "MarketPrice", [
        ("routeCode", True, "FK → Маршруты.code", None, "Код маршрута (лист «8. Маршруты»)"),
        ("vehicleTypeCode", True, "FK → Типы ТС.code", None, "Код типа ТС (лист «3. Типы ТС»)"),
        ("pricePerTrip", False, "число", None, "Рыночная цена за рейс, ₽"),
        ("pricePerPallet", False, "число", None, "Рыночная цена за паллету, ₽"),
        ("pricePerKm", False, "число", None, "Рыночная цена за км, ₽"),
        ("validFrom", True, "дата ГГГГ-ММ-ДД", None, "Действует с"),
        ("validTo", False, "дата ГГГГ-ММ-ДД", None, "Действует по"),
        ("source", False, "текст", None, "Источник цены"),
    ]),
]

wb = Workbook()

# ---------- Лист инструкции ----------
ins = wb.active
ins.title = "Инструкция"
ins.sheet_view.showGridLines = False
lines = [
    ("GrowFood Magistral — шаблон справочников", 16, True, "1F4E78"),
    ("", 10, False, None),
    ("Заполните листы и передайте файл для загрузки в БД. Один лист = один справочник.", 11, False, None),
    ("", 10, False, None),
    ("Как заполнять:", 12, True, "1F4E78"),
    ("• Заголовки столбцов менять НЕЛЬЗЯ — по ним идёт импорт.", 10, False, None),
    ("• Тёмно-синий заголовок = поле ОБЯЗАТЕЛЬНОЕ, серый = опциональное (можно оставить пустым).", 10, False, None),
    ("• Наведите курсор на заголовок столбца — во всплывающем комментарии тип, обязательность и подсказка.", 10, False, None),
    ("• Поля со списком (enum) и TRUE/FALSE — выбирайте из выпадающего списка в ячейке.", 10, False, None),
    ("• Связи (FK) указывайте КОДОМ из соответствующего листа, а не названием.", 10, False, None),
    ("  Напр. в «Контрагенты.verticalCode» пишите код из листа «1. Вертикали».", 10, False, None),
    ("• Даты — в формате ГГГГ-ММ-ДД (например 2026-01-01).", 10, False, None),
    ("• Числа — без пробелов и символа ₽ (например 85000, 18.5).", 10, False, None),
    ("", 10, False, None),
    ("Порядок заполнения (из-за связей между справочниками):", 12, True, "1F4E78"),
    ("1) Вертикали → 2) Локации → 3) Типы ТС → 4) Перевозчики → 5) Контрагенты →", 10, False, None),
    ("6) Транспорт → 7) Водители → 8) Маршруты → 9-10) Договоры → 11) Тарифы → 12) Рыночные цены.", 10, False, None),
    ("", 10, False, None),
    ("Важно про связи:", 12, True, "1F4E78"),
    ("• Транспорт ссылается на Тип ТС и Перевозчика; Маршрут — на две Локации;", 10, False, None),
    ("  Тариф — на договор (ровно один: заказчика ИЛИ перевозчика) + тип ТС; и т.д.", 10, False, None),
    ("• Сначала должен существовать код, на который ссылаетесь.", 10, False, None),
    ("• Тариф: заполняйте ЛИБО customerContractNumber, ЛИБО carrierContractNumber — строго одно из двух.", 10, False, None),
    ("", 10, False, None),
    ("Базовый сид уже есть в БД (вертикали, типы ТС, демо-контрагенты и т.п.).", 10, False, None),
    ("При загрузке строки с тем же кодом будут обновлять существующие записи.", 10, False, None),
]
r = 1
for text, size, bold, color in lines:
    c = ins.cell(row=r, column=1, value=text)
    c.font = Font(name=FONT, size=size, bold=bold, color=(color or "000000"))
    r += 1
ins.column_dimensions["A"].width = 110

# ---------- Листы справочников ----------
for title, model, cols in SHEETS:
    ws = wb.create_sheet(title=title)
    ws.sheet_view.showGridLines = False
    # заголовок-модель в строке-комментарии? Нет — только подпись через freeze
    for idx, (name, req, typ, allowed, hint) in enumerate(cols, start=1):
        col = get_column_letter(idx)
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = HDR_FONT
        cell.fill = HDR_REQ if req else HDR_OPT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        req_txt = "ОБЯЗАТЕЛЬНО" if req else "опционально"
        comment_txt = f"{req_txt}\nТип: {typ}"
        if allowed:
            comment_txt += "\nЗначения: " + ", ".join(allowed)
        comment_txt += f"\n{hint}"
        cm = Comment(comment_txt, "GrowFood")
        cm.width = 280
        cm.height = 140
        cell.comment = cm
        # ширина
        ws.column_dimensions[col].width = max(14, min(34, len(name) + 8))
        # валидация
        if allowed:
            dv = DataValidation(type="list", formula1='"' + ",".join(allowed) + '"', allow_blank=True)
            dv.error = "Выберите значение из списка"
            dv.errorTitle = "Недопустимое значение"
            ws.add_data_validation(dv)
            dv.add(f"{col}2:{col}{MAXROW}")
        # лёгкая рамка/шрифт для тела
        for rr in range(2, MAXROW + 1):
            bc = ws.cell(row=rr, column=idx)
            bc.font = BODY_FONT
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"

wb.save("Справочники_GrowFood.xlsx")
print("OK: Справочники_GrowFood.xlsx, листов:", len(wb.sheetnames))
